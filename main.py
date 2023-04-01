import openpyxl

livre=openpyxl.load_workbook("octobre.xlsx")

#pour determiner la liste des feuilles dans un fichier exel
#print(livre.sheetnames)
sheet=livre[livre.sheetnames[0]]
cells=sheet.cell(2, 4)
print(cells.value)

#print(cell.value)

#    cell = sheet.cell(row, 2)
#    print(cell.value)