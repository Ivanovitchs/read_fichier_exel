import openpyxl
readbook1=openpyxl.load_workbook("novembre.xlsx",data_only=True)
readbook2=openpyxl.load_workbook("octobre.xlsx",data_only=True)
readbook3=openpyxl.load_workbook("decembre.xlsx",data_only=True)

feuille1=readbook2.active
feuille2=readbook1.active
feuille3=readbook3.active
donne={}
listedonne=[]

def ajouterdata(feuille,d):
    for i in range(2,feuille.max_row):
        clefeuille=feuille.cell(i,1).value
        valeurfeuille=feuille.cell(i,4).value
        if not clefeuille or not valeurfeuille:
                break
        if d.get(clefeuille):
                d[clefeuille].append(valeurfeuille)
        else:
            d[clefeuille]=[valeurfeuille]

ajouterdata(feuille1,donne)
ajouterdata(feuille2,donne)
ajouterdata(feuille3,donne)

wb_sortie=openpyxl.Workbook()
feuille=wb_sortie.active

feuille["A1"]="Article"
feuille["B1"]="Octobre"
feuille["C1"]="Novembre"
feuille["D1"]="Decembre"
row=2
for i in donne.items():
    feuille.cell(row,1).value=i[0]
    valeur=i[1]
    col=2
    for j in range(len(valeur)):
        feuille.cell(row,col).value=valeur[j]
        col+=1
    row+=1

#ici on donne eune reference de la liste des données que le graphique doit contenir
#chart_ref = openpyxl.chart.Reference(sheet, min_col=2, min_row=2, max_col=sheet.max_column, max_row=2)
reference=openpyxl.chart.Reference(feuille,min_row=2,min_col=2, max_col=feuille.max_column,max_row=2)
#données la description de notre reference.
description_ref=openpyxl.chart.Series(reference, title="total Vente en frcfa")
#chart_serie = openpyxl.chart.Series(chart_ref, title="Total ventes €")
#creation de l'objet graphique
graphique=openpyxl.chart.BarChart()
graphique.append(description_ref)
graphique.title="Evolution du prix des pommes"
#on ajoute notre graphique dans la feuille
feuille.add_chart(graphique,"F2")

wb_sortie.save("totalevente.xlsx")

    
